"""
ESMA Short Position Scanner  (v2.0 — 2026-04-10)
===================================================
Downloads and parses EU/UK short position disclosures from FOUR national
regulators. Detects new positions, position changes, and crowded shorts.
Produces standardized JSON signals for the investment discovery pipeline.

Implemented Regulators (all verified accessible — April 2026):
  1. FCA  (UK)          — XLSX direct download (~580 positions)
  2. AMF  (France)      — CSV via data.gouv.fr   (~2,700 active positions)
  3. AFM  (Netherlands) — CSV export              (~15-30 positions)
  4. BaFin/Bundesanzeiger (Germany) — CSV with session cookies (~475 positions)

Not Yet Implemented (blocked from sandbox):
  5. CONSOB (Italy)     — Bot protection (Radware). Needs browser automation.
  6. CNMV  (Spain)      — Returns 403. Needs browser automation.

Signal Logic:
1. New position: entity appears at >=0.5% that wasn't in prior snapshot
2. Position increase: >=0.2% increase from prior snapshot
3. Position decrease: >=0.2% decrease (potential short covering)
4. Crowded short: 3+ entities shorting same company in current snapshot
5. Large position: any single position >=2.0%

Usage:
    python esma_short_scanner.py                    # Scan all regulators
    python esma_short_scanner.py --regulators fca amf  # Scan specific ones
    python esma_short_scanner.py --dry-run           # Print without saving
"""

import json, os, io, csv, time, logging, hashlib
from datetime import datetime, timedelta
from typing import Optional, List, Dict, Tuple
from collections import defaultdict
import requests

# Try to import from tools.mcap_cache; fall back to local import if running standalone
try:
    from tools.mcap_cache import get_market_cap_cached as _get_market_cap
except ImportError:
    from mcap_cache import get_market_cap_cached as _get_market_cap

# --- Configuration ---
FCA_URL = "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx"
AFM_URL = "https://www.afm.nl/export.aspx?type=8a46a4ef-f196-4467-a7ab-1ae1cb58f0e7&format=csv"
BAFIN_PAGE_URL = "https://www.bundesanzeiger.de/pub/en/nlp?1"
BAFIN_CSV_URL = "https://www.bundesanzeiger.de/pub/en/nlp?0--top~csv~form~panel-form-csv~resource~link"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
REQUEST_TIMEOUT = 45
ALL_REGULATORS = ["fca", "amf", "afm", "bafin"]
MARKET_CAP_FLOOR_MM = 215  # euro 200M ~ $215M
POSITION_CHANGE_THRESHOLD = 0.2
CROWDED_SHORT_MIN_HOLDERS = 3
LARGE_POSITION_THRESHOLD = 2.0
DEDUP_WINDOW_DAYS = 7
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_DIR = os.path.dirname(_SCRIPT_DIR)
SIGNALS_DIR = os.path.join(_PROJECT_DIR, "signals")
DEDUP_FILE = os.path.join(_PROJECT_DIR, "signals", "esma_dedup.json")
SNAPSHOT_DIR = os.path.join(_PROJECT_DIR, "signals", "esma_snapshots")
logger = logging.getLogger("esma_short_scanner")

REGULATOR_SOURCE_URLS = {
    "FCA": "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx",
    "AMF": "https://bdif.amf-france.org/fr?typesInformation=VAD",
    "AFM": "https://www.afm.nl/en/sector/registers/meldingenregisters/netto-shortposities-actueel",
    "BAFIN": "https://www.bundesanzeiger.de/pub/en/nlp",
}

ISIN_SUFFIX_MAP = {
    "GB": ".L", "DE": ".DE", "FR": ".PA", "NL": ".AS",
    "ES": ".MC", "IT": ".MI", "CH": ".SW", "BE": ".BR",
    "AT": ".VI", "IE": ".IR", "PT": ".LS", "SE": ".ST",
    "NO": ".OL", "DK": ".CO", "FI": ".HE", "AU": ".AX", "LU": ".PA",
}

# ===================================================================
#  REGULATOR 1: FCA (UK)
# ===================================================================
def download_fca_xlsx():
    headers = {"User-Agent": USER_AGENT}
    try:
        resp = requests.get(FCA_URL, headers=headers, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        logger.info(f"FCA: Downloaded XLSX: {len(resp.content):,} bytes")
        return resp.content
    except requests.exceptions.RequestException as e:
        logger.error(f"FCA: Failed to download XLSX: {e}")
        return None

def parse_fca_current(xlsx_bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    current_sheet = None
    for name in wb.sheetnames:
        if "current" in name.lower():
            current_sheet = wb[name]
            break
    if current_sheet is None:
        current_sheet = wb[wb.sheetnames[0]]
    positions = []
    header_seen = False
    for row in current_sheet.iter_rows(values_only=True):
        if not header_seen:
            header_seen = True
            continue
        if len(row) < 5:
            continue
        holder, issuer, isin, position_pct, position_date = row[:5]
        if not holder or not issuer:
            continue
        try:
            pct = float(position_pct)
        except (TypeError, ValueError):
            continue
        date_str = None
        if isinstance(position_date, datetime):
            date_str = position_date.strftime("%Y-%m-%d")
        elif isinstance(position_date, str):
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]:
                try:
                    date_str = datetime.strptime(position_date.strip(), fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
        positions.append({
            "regulator": "FCA", "holder_name": str(holder).strip(),
            "target_company": str(issuer).strip(),
            "isin": str(isin).strip() if isin else "",
            "position_pct": round(pct, 2), "position_date": date_str,
            "previous_position_pct": None, "change_pct": None,
            "disclosure_date": datetime.now().strftime("%Y-%m-%d"),
        })
    wb.close()
    logger.info(f"FCA: Parsed {len(positions)} current positions")
    return positions

def fetch_fca_positions():
    xlsx_bytes = download_fca_xlsx()
    if xlsx_bytes is None:
        return []
    return parse_fca_current(xlsx_bytes)

# ===================================================================
#  REGULATOR 2: AMF (France)
# ===================================================================
def _discover_amf_csv_url():
    try:
        resp = requests.get(
            "https://www.data.gouv.fr/api/1/datasets/?q=positions+courtes+nettes&page_size=5",
            timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            for dataset in data.get("data", []):
                for resource in dataset.get("resources", []):
                    url = resource.get("url", "")
                    if "amf" in url.lower() and "vad" in url.lower() and url.endswith(".csv"):
                        return url
    except Exception:
        pass
    for day_offset in range(0, 3):
        date = datetime.now() - timedelta(days=day_offset)
        ds = date.strftime("%Y%m%d")
        url = f"https://object-api.infra.data.gouv.fr/amf/VAD/export_od_vad_{ds}111500_{ds}123000.csv"
        try:
            r = requests.head(url, timeout=10)
            if r.status_code == 200:
                return url
        except Exception:
            continue
    return None

def fetch_amf_positions():
    csv_url = _discover_amf_csv_url()
    if not csv_url:
        logger.error("AMF: Could not discover CSV URL")
        return []
    try:
        resp = requests.get(csv_url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error(f"AMF: Failed to download CSV: {e}")
        return []
    text = resp.content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    try:
        next(reader)
    except StopIteration:
        return []
    active = {}
    for row in reader:
        if len(row) < 8:
            continue
        holder = row[0].strip().strip('"')
        issuer = row[2].strip().strip('"')
        ratio_str = row[3].strip().strip('"')
        isin = row[4].strip().strip('"')
        start_date = row[5].strip().strip('"')
        pub_start = row[6].strip().strip('"')
        pub_end = row[7].strip().strip('"')
        if pub_end:
            continue
        if not holder or not isin:
            continue
        try:
            pct = float(ratio_str.replace(",", "."))
        except (TypeError, ValueError):
            continue
        key = f"{holder}|{isin}"
        existing = active.get(key)
        if existing is None or start_date > existing.get("position_date", ""):
            active[key] = {
                "regulator": "AMF", "holder_name": holder,
                "target_company": issuer, "isin": isin,
                "position_pct": round(pct, 2), "position_date": start_date,
                "previous_position_pct": None, "change_pct": None,
                "disclosure_date": pub_start or datetime.now().strftime("%Y-%m-%d"),
            }
    positions = list(active.values())
    logger.info(f"AMF: Parsed {len(positions)} active positions")
    return positions

# ===================================================================
#  REGULATOR 3: AFM (Netherlands)
# ===================================================================
def fetch_afm_positions():
    headers = {"User-Agent": USER_AGENT}
    try:
        resp = requests.get(AFM_URL, headers=headers, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error(f"AFM: Failed to download CSV: {e}")
        return []
    positions = []
    text = resp.content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    try:
        next(reader)
    except StopIteration:
        return []
    for row in reader:
        if len(row) < 5:
            continue
        holder = row[0].strip().strip('"')
        issuer = row[1].strip().strip('"')
        isin = row[2].strip().strip('"')
        pct_str = row[3].strip().strip('"')
        date_raw = row[4].strip().strip('"')
        if not holder or not isin:
            continue
        try:
            pct = float(pct_str.replace(",", "."))
        except (TypeError, ValueError):
            continue
        date_str = date_raw[:10] if len(date_raw) >= 10 else date_raw
        positions.append({
            "regulator": "AFM", "holder_name": holder,
            "target_company": issuer, "isin": isin,
            "position_pct": round(pct, 2), "position_date": date_str,
            "previous_position_pct": None, "change_pct": None,
            "disclosure_date": datetime.now().strftime("%Y-%m-%d"),
        })
    logger.info(f"AFM: Parsed {len(positions)} current positions")
    return positions

# ===================================================================
#  REGULATOR 4: BaFin / Bundesanzeiger (Germany)
# ===================================================================
def fetch_bafin_positions():
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    try:
        session.get(BAFIN_PAGE_URL, timeout=REQUEST_TIMEOUT)
    except requests.exceptions.RequestException as e:
        logger.error(f"BaFin: Failed to load page: {e}")
        return []
    try:
        resp = session.get(BAFIN_CSV_URL, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.error(f"BaFin: Failed to download CSV: {e}")
        return []
    positions = []
    text = resp.content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=",")
    try:
        next(reader)
    except StopIteration:
        return []
    for row in reader:
        if len(row) < 5:
            continue
        holder = row[0].strip().strip('"')
        issuer = row[1].strip().strip('"')
        isin = row[2].strip().strip('"')
        pct_str = row[3].strip().strip('"')
        date_str = row[4].strip().strip('"')
        if not holder or not isin:
            continue
        try:
            pct = float(pct_str.replace(",", "."))
        except (TypeError, ValueError):
            continue
        positions.append({
            "regulator": "BAFIN", "holder_name": holder,
            "target_company": issuer, "isin": isin,
            "position_pct": round(pct, 2), "position_date": date_str,
            "previous_position_pct": None, "change_pct": None,
            "disclosure_date": datetime.now().strftime("%Y-%m-%d"),
        })
    logger.info(f"BaFin: Parsed {len(positions)} current positions")
    return positions

# ===================================================================
#  DISPATCH + FETCH ALL
# ===================================================================
REGULATOR_FETCHERS = {
    "fca": fetch_fca_positions, "amf": fetch_amf_positions,
    "afm": fetch_afm_positions, "bafin": fetch_bafin_positions,
}

def fetch_all_positions(regulators=None):
    targets = regulators or ALL_REGULATORS
    all_positions = []
    for reg in targets:
        fetcher = REGULATOR_FETCHERS.get(reg)
        if not fetcher:
            logger.warning(f"Unknown regulator: {reg}")
            continue
        try:
            positions = fetcher()
            all_positions.extend(positions)
            logger.info(f"{reg.upper()}: {len(positions)} positions fetched")
        except Exception as e:
            logger.error(f"{reg.upper()}: Fetch failed: {e}")
    logger.info(f"Total positions across all regulators: {len(all_positions)}")
    return all_positions

# ===================================================================
#  SNAPSHOT MANAGEMENT (multi-regulator)
# ===================================================================
def _snapshot_path(snapshot_dir, regulator, date_str):
    return os.path.join(snapshot_dir, f"{regulator}_snapshot_{date_str}.json")

def save_snapshot(positions, snapshot_dir):
    os.makedirs(snapshot_dir, exist_ok=True)
    date_str = datetime.now().strftime("%Y%m%d")
    by_reg = defaultdict(list)
    for p in positions:
        by_reg[p.get("regulator", "UNKNOWN").lower()].append(p)
    for reg, pos_list in by_reg.items():
        filepath = _snapshot_path(snapshot_dir, reg, date_str)
        with open(filepath, "w") as f:
            json.dump(pos_list, f, indent=2)
        logger.info(f"Saved snapshot: {filepath} ({len(pos_list)} positions)")

def load_prior_snapshot(snapshot_dir, regulators=None):
    if not snapshot_dir or not os.path.exists(snapshot_dir):
        return None
    targets = regulators or ALL_REGULATORS
    today = datetime.now().strftime("%Y%m%d")
    all_prior = []
    for reg in targets:
        snapshots = []
        prefix = f"{reg}_snapshot_"
        for f in os.listdir(snapshot_dir):
            if f.startswith(prefix) and f.endswith(".json"):
                date_part = f.replace(prefix, "").replace(".json", "")
                if date_part != today:
                    snapshots.append((date_part, os.path.join(snapshot_dir, f)))
        if not snapshots:
            continue
        snapshots.sort(reverse=True)
        filepath = snapshots[0][1]
        try:
            with open(filepath) as fh:
                data = json.load(fh)
            all_prior.extend(data)
            logger.info(f"Loaded prior snapshot for {reg.upper()}: {len(data)} positions")
        except Exception as e:
            logger.warning(f"Failed to load snapshot {filepath}: {e}")
    return all_prior if all_prior else None

def _build_position_index(positions):
    index = {}
    for p in positions:
        key = f"{p['holder_name']}|{p['isin']}"
        existing = index.get(key)
        if existing is None:
            index[key] = p
        else:
            if (p.get("position_date") or "") > (existing.get("position_date") or ""):
                index[key] = p
    return index

# ===================================================================
#  DIFF ENGINE
# ===================================================================
def diff_snapshots(current, prior):
    if prior is None:
        for p in current:
            p["previous_position_pct"] = None
            p["change_pct"] = None
        return current
    prior_index = _build_position_index(prior)
    for p in current:
        key = f"{p['holder_name']}|{p['isin']}"
        prev = prior_index.get(key)
        if prev:
            p["previous_position_pct"] = prev["position_pct"]
            p["change_pct"] = round(p["position_pct"] - prev["position_pct"], 2)
        else:
            p["previous_position_pct"] = None
            p["change_pct"] = None
    return current

# ===================================================================
#  SIGNAL DETECTION
# ===================================================================
def detect_crowded_shorts(positions, min_holders=CROWDED_SHORT_MIN_HOLDERS):
    by_isin = defaultdict(list)
    for p in positions:
        if p.get("isin"):
            by_isin[p["isin"]].append(p)
    return {isin: pos for isin, pos in by_isin.items() if len(pos) >= min_holders}

# DEPRECATED — use mcap_cache.get_market_cap_cached() instead
# def _get_market_cap(ticker):
#     if not ticker:
#         return None
#     try:
#         import yfinance as yf
#         stock = yf.Ticker(ticker)
#         info = stock.info
#         mcap = info.get("marketCap")
#         if mcap:
#             return mcap / 1_000_000
#     except Exception as e:
#         logger.debug(f"Market cap lookup failed for {ticker}: {e}")
#     return None

def _signal_hash(holder, isin, signal_type):
    return hashlib.md5(f"{holder}|{isin}|{signal_type}".encode()).hexdigest()

def _build_signal(position, signal_type, strength, ticker="", market_cap_mm=0, crowded_info=""):
    regulator = position.get("regulator", "UNKNOWN")
    raw_data = {
        "regulator": regulator,
        "holder_name": position.get("holder_name", ""),
        "position_pct": position.get("position_pct", 0),
        "previous_position_pct": position.get("previous_position_pct"),
        "change_pct": position.get("change_pct"),
        "position_date": position.get("position_date", ""),
        "isin": position.get("isin", ""),
    }
    if crowded_info:
        raw_data["crowded_info"] = crowded_info
    return {
        "ticker": ticker, "isin": position.get("isin", ""),
        "company_name": position.get("target_company", ""),
        "market_cap_mm": round(market_cap_mm, 1) if market_cap_mm else None,
        "signal_type": f"short_{signal_type}",
        "signal_category": "esma_short",
        "strength_estimate": strength,
        "source_url": REGULATOR_SOURCE_URLS.get(regulator, ""),
        "source_date": position.get("position_date", ""),
        "scan_date": datetime.now().strftime("%Y-%m-%d"),
        "raw_data": raw_data,
    }

def _resolve_isin_to_ticker(isin):
    if not isin or len(isin) != 12:
        return None
    country_prefix = isin[:2]
    try:
        resp = requests.post(
            "https://api.openfigi.com/v3/mapping",
            json=[{"idType": "ID_ISIN", "idValue": isin}],
            headers={"Content-Type": "application/json"}, timeout=10)
        if resp.status_code == 200:
            data = resp.json()
            if data and isinstance(data, list) and len(data) > 0:
                items = data[0].get("data", [])
                if items:
                    ticker = None
                    for item in items:
                        if item.get("securityType2") == "Common Stock":
                            ticker = item.get("ticker")
                            break
                    if not ticker:
                        ticker = items[0].get("ticker")
                    if ticker and country_prefix in ISIN_SUFFIX_MAP and "." not in ticker:
                        ticker = ticker + ISIN_SUFFIX_MAP[country_prefix]
                    return ticker
    except Exception as e:
        logger.debug(f"OpenFIGI resolve failed for {isin}: {e}")
    return None

def _load_dedup(filepath):
    if filepath and os.path.exists(filepath):
        try:
            with open(filepath) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_dedup(filepath, log):
    if filepath:
        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            with open(filepath, "w") as f:
                json.dump(log, f, indent=2)
        except Exception as e:
            logger.warning(f"Failed to save dedup: {e}")

# ===================================================================
#  MAIN SCAN
# ===================================================================
def run_scan(regulators=None, market_cap_filter=True, save_signals=True, resolve_tickers=True):
    current = fetch_all_positions(regulators)
    if not current:
        logger.warning("No positions fetched from any regulator")
        return []
    prior = load_prior_snapshot(SNAPSHOT_DIR, regulators)
    current = diff_snapshots(current, prior)
    if SNAPSHOT_DIR:
        save_snapshot(current, SNAPSHOT_DIR)
    crowded = detect_crowded_shorts(current)
    crowded_isins = set(crowded.keys())
    logger.info(f"Crowded shorts (>={CROWDED_SHORT_MIN_HOLDERS} holders): {len(crowded)} ISINs")
    dedup = _load_dedup(DEDUP_FILE)
    _ticker_cache_file = os.path.join(SIGNALS_DIR, "esma_ticker_cache.json") if SIGNALS_DIR else None
    ticker_cache = {}
    if _ticker_cache_file and os.path.exists(_ticker_cache_file):
        try:
            with open(_ticker_cache_file) as f:
                ticker_cache = json.load(f)
            logger.info(f"Loaded {len(ticker_cache)} cached ISIN->ticker mappings")
        except Exception:
            pass
    figi_call_count = 0
    FIGI_RATE_LIMIT = 25
    all_signals = []
    for pos in current:
        isin = pos.get("isin", "")
        holder = pos.get("holder_name", "")
        pct = pos.get("position_pct", 0)
        change = pos.get("change_pct")
        prev = pos.get("previous_position_pct")
        signals_for_pos = []
        if prev is None and pct >= 0.5:
            signals_for_pos.append(("new_position", 3))
        if change is not None and change >= POSITION_CHANGE_THRESHOLD:
            strength = 3 if change < 0.5 else 4
            signals_for_pos.append(("position_increase", strength))
        if change is not None and change <= -POSITION_CHANGE_THRESHOLD:
            signals_for_pos.append(("position_decrease", 2))
        if pct >= LARGE_POSITION_THRESHOLD:
            signals_for_pos.append(("large_position", 4))
        if isin in crowded_isins:
            n_holders = len(crowded[isin])
            strength = 3 if n_holders < 5 else 4
            signals_for_pos.append(("crowded_short", strength))
        if not signals_for_pos:
            continue
        signals_for_pos.sort(key=lambda x: x[1], reverse=True)
        primary_type, primary_strength = signals_for_pos[0]
        h = _signal_hash(holder, isin, primary_type)
        if h in dedup:
            try:
                first = datetime.strptime(dedup[h], "%Y-%m-%d")
                if (datetime.now() - first).days < DEDUP_WINDOW_DAYS:
                    continue
            except ValueError:
                pass
        ticker = ""
        if resolve_tickers and isin:
            if isin in ticker_cache:
                ticker = ticker_cache[isin] or ""
            elif figi_call_count < FIGI_RATE_LIMIT:
                ticker = _resolve_isin_to_ticker(isin) or ""
                ticker_cache[isin] = ticker
                figi_call_count += 1
                if figi_call_count % 5 == 0:
                    time.sleep(3)
        market_cap_mm = 0
        if market_cap_filter and ticker:
            market_cap_mm = _get_market_cap(ticker) or 0
            if 0 < market_cap_mm < MARKET_CAP_FLOOR_MM:
                logger.debug(f"Below market cap floor: {ticker} ${market_cap_mm:.0f}M")
                continue
        crowded_info = ""
        if isin in crowded_isins:
            crowded_info = json.dumps([{
                "holder": p["holder_name"], "position_pct": p["position_pct"],
                "date": p.get("position_date"), "regulator": p.get("regulator"),
            } for p in crowded[isin]])
        signal_type = "+".join(s[0] for s in signals_for_pos)
        signal = _build_signal(pos, signal_type, primary_strength,
                               ticker=ticker, market_cap_mm=market_cap_mm,
                               crowded_info=crowded_info)
        all_signals.append(signal)
        dedup[h] = datetime.now().strftime("%Y-%m-%d")
    _save_dedup(DEDUP_FILE, dedup)
    if _ticker_cache_file and ticker_cache:
        try:
            os.makedirs(os.path.dirname(_ticker_cache_file), exist_ok=True)
            with open(_ticker_cache_file, "w") as f:
                json.dump(ticker_cache, f, indent=2)
            logger.info(f"Saved {len(ticker_cache)} ISIN->ticker mappings to cache")
        except Exception as e:
            logger.warning(f"Failed to save ticker cache: {e}")
    if save_signals and SIGNALS_DIR and all_signals:
        os.makedirs(SIGNALS_DIR, exist_ok=True)
        output_file = os.path.join(SIGNALS_DIR,
            f"esma_short_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with open(output_file, "w") as f:
            json.dump(all_signals, f, indent=2)
        logger.info(f"Saved {len(all_signals)} signals to {output_file}")
    return all_signals

# ===================================================================
#  CLI
# ===================================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description="ESMA Short Position Scanner (Multi-Regulator)")
    parser.add_argument("--regulators", nargs="*", choices=ALL_REGULATORS, default=None,
                        help=f"Regulators to scan (default: all = {ALL_REGULATORS})")
    parser.add_argument("--no-market-cap", action="store_true")
    parser.add_argument("--no-resolve", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(asctime)s [%(name)s] %(levelname)s: %(message)s")
    global SIGNALS_DIR, DEDUP_FILE, SNAPSHOT_DIR
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    SIGNALS_DIR = os.path.join(project_dir, "signals")
    DEDUP_FILE = os.path.join(project_dir, "signals", "esma_short_dedup.json")
    SNAPSHOT_DIR = os.path.join(project_dir, "signals", "esma_snapshots")
    signals = run_scan(regulators=args.regulators,
                       market_cap_filter=not args.no_market_cap,
                       save_signals=not args.dry_run,
                       resolve_tickers=not args.no_resolve)
    print(f"\n{'=' * 80}")
    print(f"ESMA Short Position Scan - {len(signals)} signals found")
    print(f"{'=' * 80}")
    by_reg = defaultdict(list)
    for s in signals:
        reg = s.get("raw_data", {}).get("regulator", "?")
        by_reg[reg].append(s)
    for reg in sorted(by_reg.keys()):
        sigs = by_reg[reg]
        print(f"\n  --- {reg} ({len(sigs)} signals) ---")
        for s in sigs:
            ticker = s.get("ticker") or s.get("isin", "")[:12]
            company = s.get("company_name", "")[:25]
            holder = s.get("raw_data", {}).get("holder_name", "")[:30]
            pct = s.get("raw_data", {}).get("position_pct", 0)
            change = s.get("raw_data", {}).get("change_pct")
            stype = s.get("signal_type", "")[:35]
            strength = s.get("strength_estimate", 0)
            change_str = f"{change:+.2f}pp" if change is not None else "new"
            print(f"  [{strength}] {ticker:10s} | {company:25s} | {holder:30s} | {pct:.2f}% ({change_str}) | {stype}")

if __name__ == "__main__":
    main()
