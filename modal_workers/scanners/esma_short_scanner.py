"""
ESMA Short Position scanner — Modal port of tools/esma_short_scanner.py.

Preserved from v1 (byte-equivalent where relevant):
  - Four regulator fetchers: FCA (UK xlsx), AMF (FR csv via data.gouv.fr),
    AFM (NL csv), BaFin (DE csv with session-cookie warm).
  - Browser User-Agent per v1.
  - AMF CSV URL discovery: data.gouv.fr API first, then dated-filename HEAD
    probe fallback.
  - CSV parsing quirks preserved: AMF drops rows with pub_end (position
    closed); AFM/BaFin utf-8-sig decode; AMF/AFM/BaFin comma-decimal tolerant.
  - ISIN country-prefix → MIC mapping (expanded from v1's .suffix map so it
    works with entity_identifiers ticker_mic lookup).
  - Per-regulator daily snapshot for position-change detection.
  - Crowded-short detection: 3+ holders on same ISIN in current snapshot.
  - 7-day dedup window on (regulator, isin, holder, signal_type).

Deviations from v1:
  - No OUT_FILE / CLI; only scan(cfg).
  - No mcap_cache / market-cap floor. v1 gates positions below ~$215M mcap;
    per spec this filter is done downstream in v2 (issuer_filter + auto-caps).
  - No yfinance ticker suffixing. v1 built Yahoo-style tickers ("BARC.L");
    v2 emits the raw ISIN and lets openfigi_resolver.resolve_isin produce
    issuer_figi + ticker_local. The ISIN-prefix map is used only to derive
    the hint MIC for EntityHints, not to decorate a Yahoo ticker.
  - Storage-backed cache:
      dedup   → scanner-caches/esma/dedup.json
      snapshot→ scanner-caches/esma/snapshots/{YYYY-MM-DD}/{regulator}.json
  - source_content_hash uses "sha256:<64hex>" per spec §3.4, keyed on
    (regulator, isin, holder, position_pct, position_date).
  - signal_type narrowed to the v2 registry vocabulary:
      short_disclosure / short_buildup / short_unwind / multi_regulator_crowding.
    v1's "large_position" collapses into short_disclosure (strength boosted).
  - Budget-guard: any regulator that times out or errors is reported in
    warnings; siblings proceed. Entire-run timeout → status="partial".
  - CONSOB (IT) + CNMV (ES) remain unimplemented — same as v1.
  - OpenFIGI resolution is lazy, per-unique-ISIN, cache-backed via
    client.openfigi_cache_backend(). One call per novel ISIN; no rate-limit
    sleeping (cache + resolver's own 25-req/min limiter handle it).

IO contract:
  scan(cfg: ScannerConfig) -> ScannerResult
    - No auth required (public CSV/XLSX downloads).
    - Uses cfg.timeout_soft_s (default 60s) as wall-clock budget across all
      four regulators.
    - Applies cfg.config.top_signal_limit (default 25; 0 disables) after
      ranking crowding, projected score, strength, disclosed size, and change
      magnitude so only the top short opportunities are emitted.

Dependencies:
  - `openpyxl` required for FCA xlsx parsing. Add to modal_workers/requirements.txt.
  - `pandas` NOT required (openpyxl read_only iteration is used, mirroring v1).
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

import requests

from modal_workers.shared.dim_estimator import project_short_positioning_heuristic
from modal_workers.shared.rubric_engine import score_signal
from modal_workers.shared.scanner_base import Signal, ScannerResult
from modal_workers.shared.supabase_client import EntityHints, ScannerConfig, SupabaseClient

NAME = "esma_short_scanner"

# ---------------------------------------------------------------------------
# Constants (ported from v1)
# ---------------------------------------------------------------------------

FCA_URL = "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx"
AFM_URL = "https://www.afm.nl/export.aspx?type=8a46a4ef-f196-4467-a7ab-1ae1cb58f0e7&format=csv"
BAFIN_PAGE_URL = "https://www.bundesanzeiger.de/pub/en/nlp?1"
BAFIN_CSV_URL = "https://www.bundesanzeiger.de/pub/en/nlp?0--top~csv~form~panel-form-csv~resource~link"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
REQUEST_TIMEOUT = 45  # per-request; back to v1's value after ESMA endpoint slowness from 2026-04-24 caused 5 consecutive ReadTimeout errors. Soft budget is 1170s so 45s per request fits easily.
ALL_REGULATORS = ["fca", "amf", "afm", "bafin"]

POSITION_DISCLOSURE_THRESHOLD = 1.0  # new position must be >= this pct
POSITION_CHANGE_THRESHOLD = 0.5      # delta vs prior snapshot
LARGE_POSITION_THRESHOLD = 2.0       # bumps strength on short_disclosure
CROWDED_SHORT_MIN_HOLDERS = 6
CROWDING_MIN_TOTAL_PCT = 5.0         # sum(position_pct) across crowded holders
DEDUP_WINDOW_DAYS = 7
TOP_SIGNAL_LIMIT_DEFAULT = 25
# Drop crowding holders whose latest disclosure date is older than this.
# Pre-2026-04-27, BNP Paribas filings from 2017 and Atalan from 2025-08 were
# emitted as live signals and killed by thesis_writer as "stale". 90d covers
# the standard quarterly refile cadence — a real short refiles within that.
CROWDING_MAX_HOLDER_STALENESS_DAYS = 90

REGULATOR_SOURCE_URLS: Dict[str, str] = {
    "FCA": "https://www.fca.org.uk/publication/data/short-positions-daily-update.xlsx",
    "AMF": "https://bdif.amf-france.org/fr?typesInformation=VAD",
    "AFM": "https://www.afm.nl/en/sector/registers/meldingenregisters/netto-shortposities-actueel",
    "BAFIN": "https://www.bundesanzeiger.de/pub/en/nlp",
}

REGULATOR_COUNTRY: Dict[str, str] = {
    "FCA": "GB", "AMF": "FR", "AFM": "NL", "BAFIN": "DE",
}

_SIGNAL_TYPE_PRIORITY: Dict[str, int] = {
    "multi_regulator_crowding": 4,
    "short_unwind": 3,
    "short_buildup": 2,
    "short_disclosure": 1,
}

# ISIN country-prefix → primary MIC. Used to populate EntityHints.mic so the
# entity resolver can try ticker_mic lookup before falling back to ISIN.
# Derived from v1's ISIN_SUFFIX_MAP (Yahoo suffixes) but mapped to ISO MICs.
ISIN_MIC_MAP: Dict[str, str] = {
    "GB": "XLON", "DE": "XETR", "FR": "XPAR", "NL": "XAMS",
    "ES": "XMAD", "IT": "XMIL", "CH": "XSWX", "BE": "XBRU",
    "AT": "XWBO", "IE": "XDUB", "PT": "XLIS", "SE": "XSTO",
    "NO": "XOSL", "DK": "XCSE", "FI": "XHEL", "LU": "XLUX",
}


# ---------------------------------------------------------------------------
# REGULATOR 1: FCA (UK xlsx)
# ---------------------------------------------------------------------------

def _fetch_fca(warnings: List[str]) -> List[Dict[str, Any]]:
    try:
        resp = requests.get(FCA_URL, headers={"User-Agent": USER_AGENT},
                            timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        warnings.append(f"fca: fetch failed: {type(e).__name__}: {e}")
        return []

    try:
        import openpyxl
    except ImportError:
        warnings.append("fca: openpyxl not installed — cannot parse xlsx")
        return []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content),
                                    read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 — openpyxl raises many types
        warnings.append(f"fca: xlsx parse failed: {type(e).__name__}: {e}")
        return []

    current_sheet = None
    for name in wb.sheetnames:
        if "current" in name.lower():
            current_sheet = wb[name]
            break
    if current_sheet is None:
        current_sheet = wb[wb.sheetnames[0]]

    positions: List[Dict[str, Any]] = []
    header_seen = False
    for row in current_sheet.iter_rows(values_only=True):
        if not header_seen:
            header_seen = True
            continue
        if not row or len(row) < 5:
            continue
        holder, issuer, isin, position_pct, position_date = row[:5]
        if not holder or not issuer:
            continue
        try:
            pct = float(position_pct)
        except (TypeError, ValueError):
            continue
        date_str = _normalize_date(position_date)
        positions.append({
            "regulator": "FCA",
            "holder_name": str(holder).strip(),
            "target_company": str(issuer).strip(),
            "isin": str(isin).strip() if isin else "",
            "position_pct": round(pct, 2),
            "position_date": date_str,
        })
    wb.close()
    return positions


# ---------------------------------------------------------------------------
# REGULATOR 2: AMF (France csv via data.gouv.fr)
# ---------------------------------------------------------------------------

def _discover_amf_csv_url() -> Optional[str]:
    try:
        resp = requests.get(
            "https://www.data.gouv.fr/api/1/datasets/?q=positions+courtes+nettes&page_size=5",
            timeout=10,
        )
        if resp.status_code == 200:
            data = resp.json()
            for dataset in data.get("data", []):
                for resource in dataset.get("resources", []):
                    url = resource.get("url", "")
                    if "amf" in url.lower() and "vad" in url.lower() and url.endswith(".csv"):
                        return url
    except Exception:  # noqa: BLE001
        pass
    # Dated-filename HEAD probe fallback (same as v1).
    for day_offset in range(0, 3):
        date = datetime.now(timezone.utc) - timedelta(days=day_offset)
        ds = date.strftime("%Y%m%d")
        url = (f"https://object-api.infra.data.gouv.fr/amf/VAD/"
               f"export_od_vad_{ds}111500_{ds}123000.csv")
        try:
            r = requests.head(url, timeout=8)
            if r.status_code == 200:
                return url
        except Exception:  # noqa: BLE001
            continue
    return None


def _fetch_amf(warnings: List[str]) -> List[Dict[str, Any]]:
    csv_url = _discover_amf_csv_url()
    if not csv_url:
        warnings.append("amf: could not discover CSV URL")
        return []
    try:
        resp = requests.get(csv_url, timeout=REQUEST_TIMEOUT,
                            headers={"User-Agent": USER_AGENT})
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        warnings.append(f"amf: fetch failed: {type(e).__name__}: {e}")
        return []

    text = resp.content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    try:
        next(reader)
    except StopIteration:
        return []

    active: Dict[str, Dict[str, Any]] = {}
    for row in reader:
        if len(row) < 8:
            continue
        holder = row[0].strip().strip('"')
        issuer = row[2].strip().strip('"')
        ratio_str = row[3].strip().strip('"')
        isin = row[4].strip().strip('"')
        start_date = row[5].strip().strip('"')
        pub_end = row[7].strip().strip('"')
        # Closed positions sometimes carry whitespace-only end dates; treat
        # those the same as a real date (audit F-101, 2026-04-27).
        if pub_end and pub_end.strip():  # position closed — skip
            continue
        if not holder or not isin:
            continue
        try:
            pct = float(ratio_str.replace(",", "."))
        except (TypeError, ValueError):
            continue
        key = f"{holder}|{isin}"
        existing = active.get(key)
        if existing is None or start_date > existing.get("position_date", ""):
            active[key] = {
                "regulator": "AMF",
                "holder_name": holder,
                "target_company": issuer,
                "isin": isin,
                "position_pct": round(pct, 2),
                "position_date": start_date,
            }
    return list(active.values())


# ---------------------------------------------------------------------------
# REGULATOR 3: AFM (Netherlands csv)
# ---------------------------------------------------------------------------

def _fetch_afm(warnings: List[str]) -> List[Dict[str, Any]]:
    try:
        resp = requests.get(AFM_URL, headers={"User-Agent": USER_AGENT},
                            timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        warnings.append(f"afm: fetch failed: {type(e).__name__}: {e}")
        return []

    positions: List[Dict[str, Any]] = []
    text = resp.content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=";")
    try:
        next(reader)
    except StopIteration:
        return []

    for row in reader:
        if len(row) < 5:
            continue
        holder = row[0].strip().strip('"')
        issuer = row[1].strip().strip('"')
        isin = row[2].strip().strip('"')
        pct_str = row[3].strip().strip('"')
        date_raw = row[4].strip().strip('"')
        if not holder or not isin:
            continue
        try:
            pct = float(pct_str.replace(",", "."))
        except (TypeError, ValueError):
            continue
        date_str = date_raw[:10] if len(date_raw) >= 10 else date_raw
        positions.append({
            "regulator": "AFM",
            "holder_name": holder,
            "target_company": issuer,
            "isin": isin,
            "position_pct": round(pct, 2),
            "position_date": date_str,
        })
    return positions


# ---------------------------------------------------------------------------
# REGULATOR 4: BaFin / Bundesanzeiger (Germany csv with session cookies)
# ---------------------------------------------------------------------------

def _fetch_bafin(warnings: List[str]) -> List[Dict[str, Any]]:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    try:
        session.get(BAFIN_PAGE_URL, timeout=REQUEST_TIMEOUT)  # warm cookies
    except requests.exceptions.RequestException as e:
        warnings.append(f"bafin: page warm failed: {type(e).__name__}: {e}")
        return []
    try:
        resp = session.get(BAFIN_CSV_URL, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        warnings.append(f"bafin: csv fetch failed: {type(e).__name__}: {e}")
        return []

    positions: List[Dict[str, Any]] = []
    text = resp.content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text), delimiter=",")
    try:
        next(reader)
    except StopIteration:
        return []

    for row in reader:
        if len(row) < 5:
            continue
        holder = row[0].strip().strip('"')
        issuer = row[1].strip().strip('"')
        isin = row[2].strip().strip('"')
        pct_str = row[3].strip().strip('"')
        date_str = row[4].strip().strip('"')
        if not holder or not isin:
            continue
        try:
            pct = float(pct_str.replace(",", "."))
        except (TypeError, ValueError):
            continue
        positions.append({
            "regulator": "BAFIN",
            "holder_name": holder,
            "target_company": issuer,
            "isin": isin,
            "position_pct": round(pct, 2),
            "position_date": date_str,
        })
    return positions


_REGULATOR_FETCHERS = {
    "fca": _fetch_fca,
    "amf": _fetch_amf,
    "afm": _fetch_afm,
    "bafin": _fetch_bafin,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _normalize_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return s[:10] if len(s) >= 10 else s
    return str(value)


def _parse_source_date(date_str: Optional[str], fallback: datetime) -> datetime:
    if not date_str:
        return fallback
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(date_str, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return fallback


def _position_key(holder: str, isin: str) -> str:
    return f"{holder}|{isin}"


def _build_position_index(positions: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    index: Dict[str, Dict[str, Any]] = {}
    for p in positions:
        key = _position_key(p.get("holder_name", ""), p.get("isin", ""))
        existing = index.get(key)
        if existing is None:
            index[key] = p
        elif (p.get("position_date") or "") > (existing.get("position_date") or ""):
            index[key] = p
    return index


# Affiliate-collapse regex matches the org-suffix list used by
# insider_form4_scanner._reporter_normalized — kept in sync (audit 2026-04-27,
# F-detect-crowded). When holders' raw names differ only by entity suffix
# ("Elliott Investment Management LP" vs "Elliott Investment Management L.P."
# vs "Elliott Capital Advisors LLC"), they collapse to the same first token
# and count as ONE distinct holder, not three.
_HOLDER_ORG_INDICATORS = re.compile(
    r"\b("
    r"llc|l\.l\.c\.?|ltd|lp|l\.p\.?|inc\.?|corp\.?|company|co\.?|plc|"
    r"sa|s\.a\.?|ag|gmbh|n\.?v\.?|s\.?a\.?s\.?|trust|holdings?|partners?|"
    r"fund|funds|capital|management|advisors?|investments?|group|"
    r"americas|international|europe|asia|global|bank|asset|securities?|"
    r"associates|ventures|equities?|equity|llp|kgaa|spa|s\.r\.l\.?"
    r")\b",
    re.IGNORECASE,
)


def _normalize_holder(name: Optional[str]) -> str:
    """Canonical key for holder-affiliate dedup across regulators.

    ESMA crowding signals were inflating `holder_count` because the same fund
    files under variant names — different regulators, different entity suffix
    spellings, different affiliated funds. The DLQ bears witness: Elliott
    dual-filed FCA+AFM, Citadel filed under three affiliates, Atalan counted
    six times. Mirrors `insider_form4_scanner._reporter_normalized`.

    Org names → first-token after stripping common entity-suffix words.
    Person names (no suffix tokens, no comma) → full lowercased form.
    Empty / None → "" (callers should drop these before counting).
    """
    if not name:
        return ""
    original = name.strip()
    if not original:
        return ""
    is_org = bool(_HOLDER_ORG_INDICATORS.search(original)) or "," in original
    s = _HOLDER_ORG_INDICATORS.sub(" ", original)
    s = re.sub(r"[,\./]", " ", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    tokens = s.split()
    if not tokens:
        return original.lower()
    if is_org:
        return tokens[0]
    return " ".join(tokens)


def _is_position_fresh(position_date: Optional[str], today: datetime) -> bool:
    """True when position_date is within CROWDING_MAX_HOLDER_STALENESS_DAYS.

    Missing / unparseable dates fail-open to True — better to surface a real
    crowd with one undated holder than to silently drop a legitimate cluster.
    Operators see the row in the dashboard and can dismiss; a silent drop
    yields no signal at all.
    """
    if not position_date:
        return True
    try:
        d = datetime.strptime(position_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        return True
    return (today - d).days <= CROWDING_MAX_HOLDER_STALENESS_DAYS


def _dedup_positions(positions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    # Feeds occasionally include multiple active rows for the same (holder, isin)
    # — observed in FCA's "Current" sheet (amended filings left alongside originals).
    # signal_id is deterministic on (regulator, isin, holder, signal_type), so
    # two such rows both land in pending_signals with the same signal_id and
    # crash the bulk insert on signals_pkey. Keep the most recent position_date.
    # AMF already does this inline via its `active` dict; this helper normalizes
    # the other three regulators to the same contract.
    return list(_build_position_index(positions).values())


def _diff_vs_prior(current: List[Dict[str, Any]],
                   prior: Optional[List[Dict[str, Any]]]) -> None:
    """Annotate each current position with previous_position_pct + change_pct."""
    prior_index = _build_position_index(prior) if prior else {}
    for p in current:
        key = _position_key(p.get("holder_name", ""), p.get("isin", ""))
        prev = prior_index.get(key)
        if prev is not None:
            p["previous_position_pct"] = prev.get("position_pct")
            try:
                p["change_pct"] = round(
                    float(p["position_pct"]) - float(prev["position_pct"]), 2)
            except (TypeError, ValueError):
                p["change_pct"] = None
        else:
            p["previous_position_pct"] = None
            p["change_pct"] = None


def _detect_crowded(positions: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    today = datetime.now(timezone.utc)
    by_isin: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for p in positions:
        if not p.get("isin"):
            continue
        # Drop stale disclosures up front (90d horizon). A holder whose
        # latest filing is 6+ months old isn't part of a "current" crowd.
        # 2026-04-27 audit DLQ: BNP 2017, Atalan 2025-08, Marshall Wace
        # post-catalyst. Fail-open on missing/unparseable dates.
        if not _is_position_fresh(p.get("position_date"), today):
            continue
        by_isin[p["isin"]].append(p)

    crowded: Dict[str, List[Dict[str, Any]]] = {}
    for isin, pos in by_isin.items():
        # Count DISTINCT NORMALIZED holders, not filing rows. Same fund filed
        # under variant entity suffixes ("Elliott Investment Management LP" vs
        # "Elliott Investment Management L.P." vs cross-regulator dual-filing
        # under affiliated funds) collapse to one holder via _normalize_holder.
        # Pre-fix: fake crowds were emitted and killed downstream as
        # "single holder dual-filed" — see audit/findings_2026-04-27.md DLQ.
        unique_holders = {
            _normalize_holder(p.get("holder_name"))
            for p in pos
            if p.get("holder_name")
        }
        unique_holders.discard("")
        if len(unique_holders) < CROWDED_SHORT_MIN_HOLDERS:
            continue
        total_pct = sum((p.get("position_pct") or 0.0) for p in pos)
        if total_pct < CROWDING_MIN_TOTAL_PCT:
            continue
        crowded[isin] = pos
    return crowded


def _country_from_isin(isin: str) -> Optional[str]:
    if isin and len(isin) >= 2 and isin[:2].isalpha():
        return isin[:2].upper()
    return None


def _content_hash(regulator: str, isin: str, holder: str,
                  position_pct: float, position_date: Optional[str]) -> str:
    key = f"{regulator}|{isin}|{holder}|{position_pct}|{position_date or ''}"
    return f"sha256:{hashlib.sha256(key.encode()).hexdigest()}"


def _signal_id(regulator: str, isin: str, holder: str, signal_type: str) -> str:
    raw = f"{regulator}|{isin}|{holder}|{signal_type}"
    return f"esma_{hashlib.sha256(raw.encode()).hexdigest()[:24]}"


def _dedup_hash(regulator: str, isin: str, holder: str, signal_type: str,
                scan_date_str: str) -> str:
    # scan_date in the key: without it, a holder maintaining a 0.7% disclosure
    # day after day hashes to the same value, so the first emission blocks every
    # subsequent daily signal for 7 days. EU SSR disclosures are sticky — that
    # matched all ~2239 positions against prior dedup entries and collapsed
    # signals_emitted to 0. With scan_date in the tuple, dedup only guards
    # against double-emission within a single run (e.g. the same holder showing
    # up in two regulators' feeds for a cross-border ISIN).
    return hashlib.md5(
        f"{regulator}|{isin}|{holder}|{signal_type}|{scan_date_str}".encode()
    ).hexdigest()


# ---------------------------------------------------------------------------
# Cache IO (Supabase Storage-backed)
# ---------------------------------------------------------------------------

def _load_dedup(
    client: SupabaseClient,
    warnings: Optional[List[str]] = None,
) -> Dict[str, str]:
    raw = client.read_cache("esma", "dedup.json")
    if raw is None:
        return {}
    try:
        return json.loads(raw)
    except (ValueError, UnicodeDecodeError) as e:
        # Silent fall-through to {} would re-emit 7 days of duplicate signals
        # without any operator visibility (audit F-113, 2026-04-27). Stash the
        # corrupt blob alongside the cache for forensics, and surface a warning.
        ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
        try:
            client.write_cache(
                "esma",
                f"dedup.corrupt-{ts}.json",
                raw if isinstance(raw, (bytes, bytearray)) else str(raw).encode("utf-8"),
                content_type="application/octet-stream",
            )
        except Exception:  # noqa: BLE001 — forensics write is best-effort
            pass
        if warnings is not None:
            warnings.append(
                f"esma dedup cache parse failure ({type(e).__name__}: {e}); "
                f"resetting (corrupt copy preserved as dedup.corrupt-{ts}.json)"
            )
        return {}


def _save_dedup(client: SupabaseClient, log: Dict[str, str]) -> None:
    client.write_cache("esma", "dedup.json",
                       json.dumps(log).encode("utf-8"),
                       content_type="application/json")


def _prune_dedup(log: Dict[str, str]) -> Dict[str, str]:
    # Prune horizon must equal DEDUP_WINDOW_DAYS, not a multiple of it.
    # A longer horizon keeps stale entries past the novelty window and
    # suppresses legitimate re-emissions (audit F-111, 2026-04-27).
    cutoff = datetime.now(timezone.utc) - timedelta(days=DEDUP_WINDOW_DAYS)
    pruned: Dict[str, str] = {}
    for h, iso in log.items():
        try:
            dt = datetime.strptime(iso, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            if dt >= cutoff:
                pruned[h] = iso
        except ValueError:
            pass
    return pruned


def _snapshot_key(date_str: str, regulator: str) -> str:
    return f"snapshots/{date_str}/{regulator.lower()}.json"


def _load_prior_snapshot(client: SupabaseClient, regulator: str,
                         today_str: str) -> Optional[List[Dict[str, Any]]]:
    """Walk back up to 10 days looking for the most recent snapshot not from today."""
    for offset in range(1, 11):
        d = (datetime.strptime(today_str, "%Y-%m-%d") - timedelta(days=offset)
             ).strftime("%Y-%m-%d")
        raw = client.read_cache("esma", _snapshot_key(d, regulator))
        if raw is None:
            continue
        try:
            return json.loads(raw)
        except (ValueError, UnicodeDecodeError):
            continue
    return None


def _save_snapshot(client: SupabaseClient, regulator: str,
                   today_str: str, positions: List[Dict[str, Any]]) -> None:
    try:
        client.write_cache(
            "esma", _snapshot_key(today_str, regulator),
            json.dumps(positions).encode("utf-8"),
            content_type="application/json",
        )
    except Exception:  # noqa: BLE001 — best-effort; next run can proceed without
        pass


# ---------------------------------------------------------------------------
# Signal detection
# ---------------------------------------------------------------------------

def _classify(pos: Dict[str, Any]) -> Optional[Tuple[str, str, int]]:
    """Return (signal_type, thesis_direction, strength) or None.

    Priority (first match wins):
      1. short_unwind    — position decrease >= POSITION_CHANGE_THRESHOLD (covering = long)
      2. short_buildup   — position increase >= POSITION_CHANGE_THRESHOLD
      3. short_disclosure — new position (no prior) and >= POSITION_DISCLOSURE_THRESHOLD
    """
    pct = pos.get("position_pct") or 0.0
    prev = pos.get("previous_position_pct")
    change = pos.get("change_pct")

    if change is not None and change <= -POSITION_CHANGE_THRESHOLD:
        # Larger unwind = stronger bullish covering signal.
        strength = 3 if abs(change) >= 1.0 else 2
        return ("short_unwind", "long", strength)

    if change is not None and change >= POSITION_CHANGE_THRESHOLD:
        strength = 4 if change >= 1.0 else 3
        if pct >= LARGE_POSITION_THRESHOLD:
            strength = max(strength, 4)
        return ("short_buildup", "short", strength)

    if prev is None and pct >= POSITION_DISCLOSURE_THRESHOLD:
        strength = 4 if pct >= LARGE_POSITION_THRESHOLD else 3
        return ("short_disclosure", "short", strength)

    return None


@dataclass
class _PendingEmission:
    signal: Signal
    dedup_hash: str


def _coerce_signal_limit(value: Any, default: int) -> int:
    if value is None or isinstance(value, bool):
        return default
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed >= 0 else default


def _project_short_score(sig: Signal) -> float:
    """Internal heuristic ranking score for top_signal_limit selection.

    Calls the preserved short_positioning heuristic directly (the public
    `estimate_dimensions("short_positioning", ...)` returns None now —
    short_positioning emits unscored and is AI-resolved). Score here is
    used only to rank pending emissions; it is NOT persisted.
    """
    estimate = project_short_positioning_heuristic(sig.raw_payload)
    if estimate is None:
        return 0.0
    scored = score_signal({
        "scoring_profile": "short_positioning",
        "raw_data": {"dimensions": estimate.dimensions},
    }, provenance="heuristic")
    return float(scored.get("score") or 0.0)


def _signal_priority(sig: Signal) -> Tuple[int, float, int, float, float, str, str]:
    raw = sig.raw_payload
    disclosed = raw.get("total_disclosed_pct")
    if not isinstance(disclosed, (int, float)):
        disclosed = raw.get("position_pct")
    if not isinstance(disclosed, (int, float)):
        disclosed = 0.0

    change = raw.get("change_pct")
    delta = abs(change) if isinstance(change, (int, float)) else 0.0

    return (
        _SIGNAL_TYPE_PRIORITY.get(sig.signal_type, 0),
        _project_short_score(sig),
        sig.strength_estimate or 0,
        float(disclosed),
        float(delta),
        sig.source_date.isoformat(),
        sig.signal_id,
    )


def _apply_top_signal_limit(
    pending: List[_PendingEmission],
    limit: int,
) -> Tuple[List[_PendingEmission], List[_PendingEmission]]:
    if limit == 0 or len(pending) <= limit:
        return pending, []
    ranked = sorted(pending, key=lambda item: _signal_priority(item.signal), reverse=True)
    return ranked[:limit], ranked[limit:]


def _signal_type_breakdown(pending: List[_PendingEmission]) -> str:
    counts: Dict[str, int] = defaultdict(int)
    for item in pending:
        counts[item.signal.signal_type] += 1
    return ", ".join(f"{signal_type}={count}" for signal_type, count in sorted(counts.items()))


def _build_signal(pos: Dict[str, Any], signal_type: str, direction: str,
                  strength: int, scan_date: datetime,
                  issuer_figi: Optional[str],
                  ticker_local: Optional[str] = None) -> Signal:
    regulator = pos.get("regulator", "UNKNOWN")
    isin = pos.get("isin", "")
    holder = pos.get("holder_name", "")
    position_pct = pos.get("position_pct", 0.0)
    position_date = pos.get("position_date")
    country = _country_from_isin(isin)
    mic = ISIN_MIC_MAP.get(country) if country else None

    source_content_hash = _content_hash(
        regulator, isin, holder, position_pct, position_date)
    signal_id = _signal_id(regulator, isin, holder, signal_type)
    source_date = _parse_source_date(position_date, scan_date)

    raw_payload: Dict[str, Any] = {
        "regulator": regulator,
        "holder_name": holder,
        "target_company": pos.get("target_company", ""),
        "isin": isin,
        "position_pct": position_pct,
        "previous_position_pct": pos.get("previous_position_pct"),
        "change_pct": pos.get("change_pct"),
        "position_date": position_date,
    }

    entity_hints = EntityHints(
        issuer_figi=issuer_figi,
        ticker=ticker_local,
        isin=isin or None,
        mic=mic,
        name=pos.get("target_company") or None,
        country=country,
    )

    return Signal(
        signal_id=signal_id,
        source_content_hash=source_content_hash,
        source_date=source_date,
        scan_date=scan_date,
        signal_type=signal_type,
        raw_payload=raw_payload,
        source_url=REGULATOR_SOURCE_URLS.get(regulator),
        issuer_figi=issuer_figi,
        entity_hints=entity_hints,
        thesis_direction=direction,
        strength_estimate=strength,
    )


def _build_crowding_signal(isin: str, positions: List[Dict[str, Any]],
                           scan_date: datetime,
                           issuer_figi: Optional[str],
                           ticker_local: Optional[str] = None) -> Signal:
    regulators = sorted({p.get("regulator", "") for p in positions if p.get("regulator")})
    holders = [{
        "regulator": p.get("regulator"),
        "holder_name": p.get("holder_name"),
        "position_pct": p.get("position_pct"),
        "position_date": p.get("position_date"),
    } for p in positions]
    total_pct = sum((p.get("position_pct") or 0.0) for p in positions)
    target_company = next(
        (p.get("target_company") for p in positions if p.get("target_company")), "")
    country = _country_from_isin(isin)
    mic = ISIN_MIC_MAP.get(country) if country else None

    # Content hash keyed on the aggregate membership so hash changes only when
    # holders join/leave, not when any single pct ticks.
    membership_key = "|".join(
        sorted(f"{p.get('regulator')}:{p.get('holder_name')}:{p.get('position_pct')}"
               for p in positions))
    source_content_hash = (
        f"sha256:{hashlib.sha256(f'CROWDING|{isin}|{membership_key}'.encode()).hexdigest()}"
    )
    signal_id = f"esma_{hashlib.sha256(f'CROWDING|{isin}|{membership_key}'.encode()).hexdigest()[:24]}"

    # Count distinct *normalized* holders, not filing rows. `len(positions)`
    # over-reported when one fund filed under variant names or dual-filed
    # across regulators — downstream thesis_writer killed those as fake
    # crowding (audit/findings_2026-04-27.md DLQ).
    unique_holder_count = len({
        _normalize_holder(p.get("holder_name"))
        for p in positions
        if p.get("holder_name")
    } - {""})

    strength = 4 if unique_holder_count >= 5 else 3
    if len(regulators) >= 3:  # cross-regulator crowding is the strongest signal
        strength = max(strength, 4)

    raw_payload: Dict[str, Any] = {
        "isin": isin,
        "target_company": target_company,
        "holder_count": unique_holder_count,
        "filing_row_count": len(positions),
        "regulators": regulators,
        "total_disclosed_pct": round(total_pct, 2),
        "holders": holders,
    }

    entity_hints = EntityHints(
        issuer_figi=issuer_figi,
        ticker=ticker_local,
        isin=isin or None,
        mic=mic,
        name=target_company or None,
        country=country,
    )

    return Signal(
        signal_id=signal_id,
        source_content_hash=source_content_hash,
        source_date=scan_date,
        scan_date=scan_date,
        signal_type="multi_regulator_crowding",
        raw_payload=raw_payload,
        source_url=None,
        issuer_figi=issuer_figi,
        entity_hints=entity_hints,
        thesis_direction="neutral",
        strength_estimate=strength,
    )


# ---------------------------------------------------------------------------
# scan entrypoint
# ---------------------------------------------------------------------------

def scan(cfg: ScannerConfig) -> ScannerResult:
    client = SupabaseClient()

    # Wire openfigi cache through Supabase Storage (cheap ISIN lookups).
    try:
        from modal_workers.shared.openfigi_resolver import set_cache_backend
        set_cache_backend(*client.openfigi_cache_backend())
    except Exception:  # noqa: BLE001 — resolver still works with file backend
        pass

    scan_date = datetime.now(timezone.utc)
    today_str = scan_date.strftime("%Y-%m-%d")
    budget = max(20, cfg.timeout_soft_s - 5)  # leave headroom for cache writes
    scan_start = time.time()

    warnings: List[str] = []
    pending_signals: List[_PendingEmission] = []
    fetched_total = 0

    # --- Fetch all regulators concurrently with per-regulator hard cap ---
    # v1-era sequential loop would let one stalling fetcher (observed: BaFin or
    # AMF discovery) burn the full budget. Now: 4 threads in parallel; each
    # bounded by `per_reg_cap` via future.result(timeout=). A single hang no
    # longer blocks the others — we harvest partial results and flag the stall
    # in warnings.
    #
    # CRITICAL: the ThreadPoolExecutor is NOT used as a `with` context manager
    # because its default __exit__ calls shutdown(wait=True) and blocks until
    # every submitted task finishes. A stuck fetcher (TCP connection half-open
    # on a misbehaving regulator endpoint) would then hold the scanner at
    # shutdown and blow the Modal hard-timeout. Instead we drive the executor
    # manually and call shutdown(wait=False, cancel_futures=True) at the end:
    # pending tasks are cancelled, running-but-stuck tasks are abandoned (daemon
    # threads will die with the process on Modal container exit).
    per_reg_cap = min(budget, 90)

    def _run_regulator(reg: str) -> Tuple[str, List[Dict[str, Any]], List[str]]:
        local_warnings: List[str] = []
        fetcher = _REGULATOR_FETCHERS[reg]
        try:
            positions = fetcher(local_warnings)
            return reg, positions, local_warnings
        except Exception as e:  # noqa: BLE001
            local_warnings.append(f"{reg}: fetch raised {type(e).__name__}: {e}")
            return reg, [], local_warnings

    current_by_regulator: Dict[str, List[Dict[str, Any]]] = {}
    ex = ThreadPoolExecutor(max_workers=len(ALL_REGULATORS))
    try:
        futures = {ex.submit(_run_regulator, reg): reg for reg in ALL_REGULATORS}
        for future, reg in futures.items():
            remaining = budget - (time.time() - scan_start)
            if remaining <= 0:
                warnings.append(f"{reg}: skipped — wall-clock budget ({budget}s) exceeded")
                future.cancel()
                continue
            try:
                _reg, positions, local_warnings = future.result(timeout=min(per_reg_cap, remaining))
            except FuturesTimeout:
                warnings.append(
                    f"{reg}: per-regulator timeout ({per_reg_cap}s); abandoning thread, siblings harvested")
                continue
            except Exception as e:  # noqa: BLE001
                warnings.append(f"{reg}: unexpected wrapper exception {type(e).__name__}: {e}")
                continue
            warnings.extend(local_warnings)
            current_by_regulator[reg] = positions
            fetched_total += len(positions)
    finally:
        # Don't wait for stuck fetchers. cancel_futures requires 3.9+; we rely on it.
        ex.shutdown(wait=False, cancel_futures=True)

    if not any(current_by_regulator.values()):
        status = "error" if warnings else "ok"
        return ScannerResult(
            scanner=NAME, status=status, signals=[],
            warnings=warnings, fetched_records=fetched_total,
        )

    # --- Diff each regulator vs its own prior snapshot ---
    all_current: List[Dict[str, Any]] = []
    for reg in list(current_by_regulator.keys()):
        positions = _dedup_positions(current_by_regulator[reg])
        current_by_regulator[reg] = positions  # snapshot save uses deduped list
        prior = _load_prior_snapshot(client, reg.upper(), today_str)
        _diff_vs_prior(positions, prior)
        all_current.extend(positions)

    # --- Multi-regulator crowding detection (union across regulators) ---
    crowded = _detect_crowded(all_current)

    # --- Dedup log ---
    dedup_log = _prune_dedup(_load_dedup(client, warnings))
    novel_window = timedelta(days=DEDUP_WINDOW_DAYS)

    # --- Batch OpenFIGI resolution: one API call for all unique ISINs ---
    # v1-era pattern: per-position resolve_isin() = N serial calls, throttled at
    # 25 req/min anonymous (OPENFIGI_API_KEY unset in current deploy). At ~80
    # positions that's ~200s of rate-limited waiting — enough to exhaust the
    # 240s Modal hard-timeout after the fetch phase consumed ~40s. Batching
    # sends one _post_batch() request of up to 100 items, resolving the whole
    # run in a single roundtrip (~1s). resolve_batch() already handles cache
    # hits per-ISIN so re-runs are free.
    unique_isins: List[str] = sorted({
        pos.get("isin", "") for pos in all_current
        if isinstance(pos.get("isin"), str) and len(pos.get("isin", "")) == 12
    })
    figi_by_isin: Dict[str, Optional[str]] = {}
    ticker_by_isin: Dict[str, Optional[str]] = {}
    if unique_isins:
        try:
            from modal_workers.shared.openfigi_resolver import resolve_batch
            queries = [{"idType": "ID_ISIN", "idValue": isin} for isin in unique_isins]
            results = resolve_batch(queries)
            # resolve_batch returns resolutions in the same order as queries (ignoring
            # dropped Nones, which shouldn't happen on valid ISIN input).
            for isin, res in zip(unique_isins, results):
                figi_by_isin[isin] = res.issuer_figi if res.resolved else None
                # ticker_local feeds EntityHints.ticker so entities.primary_ticker is
                # populated on entity creation; without it the dashboard signals table
                # renders "?" because its tile reads primary_ticker.
                ticker_by_isin[isin] = res.ticker_local if res.resolved else None
        except Exception as e:  # noqa: BLE001
            warnings.append(f"openfigi batch resolve failed: {type(e).__name__}: {e}")

    def _figi_for(isin: str) -> Optional[str]:
        if not isin or len(isin) != 12:
            return None
        return figi_by_isin.get(isin)

    def _ticker_for(isin: str) -> Optional[str]:
        if not isin or len(isin) != 12:
            return None
        return ticker_by_isin.get(isin)

    # --- Emit per-position signals ---
    # Drop-reason telemetry — counts every position that didn't become a signal
    # so we can diagnose "2239 records fetched, 0 signals emitted" at a glance
    # instead of reaching for Modal stdout. Exposed as a warnings[] line at run end.
    drop: Dict[str, int] = {
        "classified_none": 0,
        "classified_none_low_pct": 0,      # prev=None and pct < 0.5 — sub-threshold disclosure
        "classified_none_no_change": 0,    # prev set, change < threshold
        "classified_none_missing_pct": 0,  # position_pct missing or not parseable
        "dedup_suppressed": 0,
        "top_signal_capped": 0,
    }

    for pos in all_current:
        if time.time() - scan_start > budget:
            warnings.append("signal-build loop aborted — budget exceeded")
            break
        classified = _classify(pos)
        if classified is None:
            drop["classified_none"] += 1
            pct_val = pos.get("position_pct")
            prev_val = pos.get("previous_position_pct")
            change_val = pos.get("change_pct")
            if pct_val is None:
                drop["classified_none_missing_pct"] += 1
            elif prev_val is None and pct_val < POSITION_DISCLOSURE_THRESHOLD:
                drop["classified_none_low_pct"] += 1
            elif change_val is not None and abs(change_val) < POSITION_CHANGE_THRESHOLD:
                drop["classified_none_no_change"] += 1
            continue
        signal_type, direction, strength = classified

        regulator = pos.get("regulator", "UNKNOWN")
        isin = pos.get("isin", "")
        holder = pos.get("holder_name", "")

        # Dedup on (regulator, isin, holder, signal_type, scan_date): prevents
        # double-emission within the same run, not across days.
        h = _dedup_hash(regulator, isin, holder, signal_type, today_str)
        prior_iso = dedup_log.get(h)
        if prior_iso:
            try:
                first = datetime.strptime(prior_iso, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if (scan_date - first) < novel_window:
                    drop["dedup_suppressed"] += 1
                    continue
            except ValueError:
                pass

        issuer_figi = _figi_for(isin)
        ticker_local = _ticker_for(isin)
        sig = _build_signal(pos, signal_type, direction, strength, scan_date,
                            issuer_figi, ticker_local)
        pending_signals.append(_PendingEmission(signal=sig, dedup_hash=h))

    if any(v > 0 for v in drop.values()):
        warnings.append(
            "drop_reasons: " + ", ".join(f"{k}={v}" for k, v in drop.items() if v > 0))

    # --- Emit crowding signals (one per crowded ISIN) ---
    for isin, positions in crowded.items():
        if time.time() - scan_start > budget:
            warnings.append("crowding loop aborted — budget exceeded")
            break
        h = _dedup_hash("CROWDING", isin, "", "multi_regulator_crowding", today_str)
        prior_iso = dedup_log.get(h)
        if prior_iso:
            try:
                first = datetime.strptime(prior_iso, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                if (scan_date - first) < novel_window:
                    continue
            except ValueError:
                pass
        issuer_figi = _figi_for(isin)
        ticker_local = _ticker_for(isin)
        pending_signals.append(_PendingEmission(
            signal=_build_crowding_signal(isin, positions, scan_date, issuer_figi, ticker_local),
            dedup_hash=h,
        ))

    # Defense-in-depth: the main loop trusts dedup_log to prevent duplicate
    # signal_ids within a run, but the log is only written AFTER top-cap, so
    # two positions sharing (regulator, isin, holder, signal_type) both slip
    # through and collide on signals_pkey at bulk insert (the ON CONFLICT
    # clause targets source_content_hash, not signal_id). Per-regulator
    # _dedup_positions above should prevent this; this is a belt-and-braces
    # guard so a future feed quirk can't crash the whole batch.
    seen_signal_ids: set[str] = set()
    unique_pending: List[_PendingEmission] = []
    collided = 0
    for item in pending_signals:
        if item.signal.signal_id in seen_signal_ids:
            collided += 1
            continue
        seen_signal_ids.add(item.signal.signal_id)
        unique_pending.append(item)
    if collided:
        warnings.append(f"signal_id_collision: dropped {collided} duplicate emission(s)")
    pending_signals = unique_pending

    top_signal_limit = _coerce_signal_limit(
        cfg.config.get("top_signal_limit"),
        TOP_SIGNAL_LIMIT_DEFAULT,
    )
    kept_emissions, dropped_emissions = _apply_top_signal_limit(
        pending_signals,
        top_signal_limit,
    )
    if dropped_emissions:
        drop["top_signal_capped"] = len(dropped_emissions)
        warnings.append(
            f"top_signal_limit: kept {len(kept_emissions)}/{len(pending_signals)} "
            f"(limit={top_signal_limit}); dropped_by_type: {_signal_type_breakdown(dropped_emissions)}"
        )

    for item in kept_emissions:
        hints = item.signal.entity_hints
        ticker = hints.ticker if hints else None
        mic = hints.mic if hints else None
        if not ticker:
            continue
        try:
            from modal_workers.shared.market_snapshot import load_market_snapshot
            snapshot = load_market_snapshot(ticker, mic=mic, client=client)
            if snapshot:
                item.signal.raw_payload.update(snapshot)
        except Exception as e:  # noqa: BLE001 — best-effort enrichment only
            from modal_workers.observability import record_snapshot_fetch_failure
            record_snapshot_fetch_failure(client, scanner_name="esma_short_scanner", ticker=ticker, exc=e)
            continue

    signals = [item.signal for item in kept_emissions]
    for item in kept_emissions:
        dedup_log[item.dedup_hash] = today_str

    # --- Persist snapshot (per regulator) + dedup — DEFERRED to post-insert. ---
    # Before 2026-04-21 these ran here, before scanner_base inserted the signals.
    # A Modal kill between scan() return and insert_signals() would then save
    # cache state without any matching rows in the signals table — the classic
    # dedup-poisoning incident (memory: v2_migration_state.md). The after_insert
    # hook runs ONLY if insert_signals didn't raise.
    _snapshot_captures = {reg.upper(): positions for reg, positions in current_by_regulator.items()}
    _dedup_capture = dedup_log

    def _persist_after_insert() -> None:
        for reg, positions in _snapshot_captures.items():
            try:
                _save_snapshot(client, reg, today_str, positions)
            except Exception:  # noqa: BLE001 — per-regulator failure is advisory
                pass
        try:
            _save_dedup(client, _dedup_capture)
        except Exception:  # noqa: BLE001 — dedup save failure is advisory here; next run
            pass                 # starts from a slightly older log, worst case re-emits a dup

    status = "partial" if warnings else "ok"
    return ScannerResult(
        scanner=NAME,
        status=status,
        signals=signals,
        warnings=warnings,
        fetched_records=fetched_total,
        after_insert=_persist_after_insert,
    )
